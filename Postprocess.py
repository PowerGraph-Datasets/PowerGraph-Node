import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

# Specify the path to your Excel file
models_path = 'C:\\Users\\avarbella\\OneDrive - ETH Zurich\\Documents\\01_GraphGym\\PowerGraph-masterdataset-node\\code\\modelbig\\'
rnd_seeds = [0, 100, 300, 700, 1000]
tasks = ['nodeopf']
powergrids = ['ieee24', 'ieee39', 'uk', 'ieee118']
n_bus = [24, 39, 29, 118]
models = ['gcn', 'gin', 'gat', 'transformer']
# Specify the sheet name
sheet_name = 'Metrics'
sheet_name_d = 'Data'

sheet_node = {
    'Power grid': ['ieee24', '', '', '', 'ieee39', '', '', '', 'uk', '', '', '', 'ieee118', '', '', ''],
    'MPL type': ['gcn', 'gin', 'gat', 'transformer', 'gcn', 'gin', 'gat', 'transformer', 'gcn', 'gin', 'gat', 'transformer', 'gcn', 'gin', 'gat', 'transformer'],
    'mse': [],
    'rmse': [],
}

#sheet_node = {
#    'Power grid': ['ieee24', '', '', 'ieee39', '', '', 'uk', '', ''],
#    'MPL type': ['gcn', 'gin', 'gat', 'gcn', 'gin', 'gat', 'gcn', 'gin', 'gat'],
#    'mse': [],
#    'rmse': [],
#}


results = {}
smallest_mse = float('inf')  # Initialize smallest MSE to positive infinity
best_model = None
best_random_seed = None

for i,powergrid in enumerate(powergrids):
    results[powergrid] = {}
    for model in models:
        results[powergrid][model] = {}
        for task in tasks:
            results[powergrid][model][task] = {}
            if task == tasks[0]:
                metrics = ['mse', 'rmse']
                results[powergrid][model][task][metrics[0]] = {}
                results[powergrid][model][task][metrics[1]] = {}
                mse = []
                rmsescore = []
            for rnd_seed in rnd_seeds:
                specific_excel_file_path = models_path + powergrid + '\\' + 'summary' + powergrid + '_' + model + '_' + task + '_3l_16h_' + str(rnd_seed) +'s' + '.xlsx'
                print(specific_excel_file_path)
                workbook = openpyxl.load_workbook(specific_excel_file_path)
                sheet = workbook[sheet_name]
                if task == tasks[0]:
                    mse_value = sheet['B2'].value
                    mse.append(mse_value)
                    rmsescore.append(sheet['B3'].value)
                    if mse_value < smallest_mse:
                        smallest_mse = mse_value
                        best_model = model
                        best_random_seed = rnd_seed
                workbook.close()

            if task == tasks[0]:
                sheet_node[metrics[0]].append(str(np.format_float_scientific(np.mean(mse), precision=4))+'±'+str(np.format_float_scientific(np.std(mse), precision=4)))
                sheet_node[metrics[1]].append(str(np.format_float_scientific(np.mean(rmsescore), precision=4))+'±'+str(np.format_float_scientific(np.std(rmsescore), precision=4)))
    if task == 'node':
        plt.figure(i)
        best_excel_file_path = models_path + powergrid + '\\' + 'summary' + powergrid + '_' + best_model + '_' + task + '_3l_16h_' + str(
        best_random_seed) + 's' + '.xlsx'
        workbookplot = openpyxl.load_workbook(best_excel_file_path)
        sheet = workbookplot[sheet_name_d]
        # Get data from columns A, B, C, and D
        v_targ = [cell.value for cell in sheet['A']]
        t_targ = [cell.value for cell in sheet['B']]
        pg_targ = [cell.value for cell in sheet['C']]
        qg_targ = [cell.value for cell in sheet['D']]
        v_pred = [cell.value for cell in sheet['E']]
        t_pred = [cell.value for cell in sheet['F']]
        pg_pred = [cell.value for cell in sheet['G']]
        qg_pred = [cell.value for cell in sheet['H']]

        mse_ac = (sum(np.abs(np.array(v_targ[1:n_bus[i]]) - np.array(v_pred[1:n_bus[i]]))))
        mse_bd = (sum(np.abs(np.array(t_targ[1:n_bus[i]]) - np.array(t_pred[1:n_bus[i]]))))
        mse_pg = sum(np.abs(np.array(pg_targ[1:n_bus[i]] - np.array(pg_pred[1:n_bus[i]]))))
        mse_qg = sum(np.abs(np.array(qg_targ[1:n_bus[i]] - np.array(qg_pred[1:n_bus[i]]))))
        plt.bar(['V [p.u.]', 'T [degree]', 'Pg [MW]', 'Qg [MVAr]'], [mse_ac, mse_bd, mse_pg, mse_qg])
        plt.xlabel('quantity')
        plt.ylabel('Mean Absolute Error (MAE)')
        plt.title(f'error total test set {powergrid}')
        plt.savefig(f'mse_bar_plot_{powergrid}.png')


    elif task == 'nodeopf':
        plt.figure(i)
        best_excel_file_path = models_path + powergrid + '\\' + 'summary' + powergrid + '_' + best_model + '_' + task + '_3l_16h_' + str(
        best_random_seed) + 's' + '.xlsx'
        workbookplot = openpyxl.load_workbook(best_excel_file_path)
        sheet = workbookplot[sheet_name_d]
        # Get data from columns A, B, C, and D
        v_targ = [cell.value for cell in sheet['A']]
        t_targ = [cell.value for cell in sheet['B']]
        pg_targ = [cell.value for cell in sheet['C']]
        qg_targ = [cell.value for cell in sheet['D']]
        v_pred = [cell.value for cell in sheet['E']]
        t_pred = [cell.value for cell in sheet['F']]
        pg_pred = [cell.value for cell in sheet['G']]
        qg_pred = [cell.value for cell in sheet['H']]

        # Calculate MSE between columns
        mse_ac = sum(np.abs(np.array(v_targ[1:n_bus[i]] - np.array(v_pred[1:n_bus[i]]))))
        mse_bd = sum(np.abs(np.array(t_targ[1:n_bus[i]] - np.array(t_pred[1:n_bus[i]]))))
        mse_pg = sum(np.abs(np.array(pg_targ[1:n_bus[i]] - np.array(pg_pred[1:n_bus[i]]))))
        mse_qg = sum(np.abs(np.array(qg_targ[1:n_bus[i]] - np.array(qg_pred[1:n_bus[i]]))))
        print(f"GRID {powergrid} v:{mse_ac},t:{mse_bd}, pg:{mse_pg},qg:{mse_qg}")

        plt.bar(['V [p.u.]', 'T [degree]', 'Pg [MW]', 'Qg [MVAr]'], [mse_ac, mse_bd, mse_pg, mse_qg])
        plt.xlabel('quantity')
        plt.ylabel('Mean Absolute Error (MAE)')
        plt.title(f'error total test set {powergrid}')
        plt.savefig(f'mse_bar_plot_{powergrid}.png')


df_sheet_Regression = pd.DataFrame(sheet_node)


excel_file_path = f'processed_results_{task}_modelbig.xlsx'

# Create a Pandas Excel writer using XlsxWriter as the engine
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    # Write each DataFrame to a different sheet
    df_sheet_Regression.to_excel(writer, sheet_name='node', index=False)
    for powergrid in powergrids:
        worksheet = writer.sheets['node']
        worksheet.insert_image('E1', f'mse_bar_plot_{powergrid}.png')

"""
# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Select the desired sheet
sheet = workbook[sheet_name]

# Specify the cell or range of cells you want to retrieve
cell_value = sheet['A1'].value  # Change 'A1' to your desired cell reference

# Or, if you want to iterate through a range of cells, for example, column A from row 1 to 10
column_values = [sheet[f'A{i}'].value for i in range(1, 11)]

# Print the results
print(f"Value in A1: {cell_value}")
print(f"Values in column A from row 1 to 10: {column_values}")
"""
# Close the workbook when done
workbook.close()
