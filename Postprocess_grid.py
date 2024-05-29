import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

# Specify the path to your Excel file
models_path = 'path'# Specify the path to the folder containing the Excel files
rnd_seeds = [0, 100, 300, 700, 1000]
hyperparameters = ['1l_8h','2l_8h','3l_8h','1l_16h','2l_16h','3l_16h','1l_32h','2l_32h','3l_32h']
tasks = ['node','nodeopf']
powergrids = ['uk']
n_bus = 29
models = ['gcn', 'gin', 'gat', 'transformer']
# Specify the sheet name
sheet_name = 'Metrics'
sheet_name_d = 'Data'

sheet_node = {
    'hyperparameters': ['1l_8h','','','','2l_8h','','','','3l_8h','','','','1l_16h','','','','2l_16h','','','','3l_16h','','','','1l_32h','','','','2l_32h','','','','3l_32h','','',''],
    'MPL type': ['gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer'],
    'mse': [],
    'rmse': [],
}
sheet_nodeopf = {
    'hyperparameters': ['1l_8h','','','','2l_8h','','','','3l_8h','','','','1l_16h','','','','2l_16h','','','','3l_16h','','','','1l_32h','','','','2l_32h','','','','3l_32h','','',''],
    'MPL type': ['gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer','gcn', 'gin', 'gat', 'transformer'],
    'mse': [],
    'rmse': [],
}
results = {}
resultsopf = {}
smallest_mse = float('inf')  # Initialize smallest MSE to positive infinity
smallest_mseopf = float('inf')  # Initialize smallest MSE to positive infinity
best_model = None
best_modelopf = None
best_random_seed = None
best_random_seedopf = None
best_hp = None
best_hpopf = None

for task in tasks:
    for layer in hyperparameters:
        results[layer] = {}
        resultsopf[layer] = {}
        powergrid = powergrids[0]
        for model in models:
            results[layer][model] = {}
            resultsopf[layer][model] = {}
            if task == 'node':
                metrics = ['mse', 'rmse']
                results[layer][model][metrics[0]] = {}
                results[layer][model][metrics[1]] = {}
                mse = []
                rmsescore = []
            if task == 'nodeopf':
                metrics = ['mse', 'rmse']
                resultsopf[layer][model][metrics[0]] = {}
                resultsopf[layer][model][metrics[1]] = {}
                mse = []
                rmsescore = []

            for rnd_seed in rnd_seeds:
                specific_excel_file_path = models_path + powergrid + '\\' + 'summary' + powergrid + '_' + model + '_' + task + '_'+layer+'_' + str(rnd_seed) +'s' + '.xlsx'
                print(specific_excel_file_path)
                workbook = openpyxl.load_workbook(specific_excel_file_path)
                sheet = workbook[sheet_name]
                if task == 'node':
                    mse_value = sheet['B2'].value
                    mse.append(mse_value)
                    rmsescore.append(sheet['B3'].value)
                    if mse_value < smallest_mse:
                        smallest_mse = mse_value
                        best_model = model
                        best_random_seed = rnd_seed
                        best_hp = layer

                if task == 'nodeopf':
                    mse_value = sheet['B2'].value
                    mse.append(mse_value)
                    rmsescore.append(sheet['B3'].value)
                    if mse_value < smallest_mseopf:
                        smallest_mseopf = mse_value
                        best_modelopf = model
                        best_random_seedopf = rnd_seed
                        best_hpopf = layer
                workbook.close()

            if task == 'node':
                sheet_node[metrics[0]].append(str(np.format_float_scientific(np.mean(mse), precision=4))+'±'+str(np.format_float_scientific(np.std(mse), precision=4)))
                sheet_node[metrics[1]].append(str(np.format_float_scientific(np.mean(rmsescore), precision=4))+'±'+str(np.format_float_scientific(np.std(rmsescore), precision=4)))
            if task == 'nodeopf':
                sheet_nodeopf[metrics[0]].append(str(np.format_float_scientific(np.mean(mse), precision=4))+'±'+str(np.format_float_scientific(np.std(mse), precision=4)))
                sheet_nodeopf[metrics[1]].append(str(np.format_float_scientific(np.mean(rmsescore), precision=4))+'±'+str(np.format_float_scientific(np.std(rmsescore), precision=4)))

    if task == 'node':
        plt.figure(0)
        best_excel_file_path = models_path + powergrid + '\\' + 'summary' + powergrid + '_' + best_model + '_' + task + '_'+best_hp+'_' + str(
        best_random_seed) + 's' + '.xlsx'
        workbookplot = openpyxl.load_workbook(best_excel_file_path)
        sheet = workbookplot[sheet_name_d]
    # Get data from columns A, B, C, and D
        v_targ = [cell.value for cell in sheet['A']]
        t_targ = [cell.value for cell in sheet['B']]
        pg_targ = [cell.value for cell in sheet['C']]
        qg_targ = [cell.value for cell in sheet['D']]
        v_pred = [cell.value for cell in sheet['E']]
        t_pred = [cell.value for cell in sheet['F']]
        pg_pred = [cell.value for cell in sheet['G']]
        qg_pred = [cell.value for cell in sheet['H']]

        mse_ac = (sum(np.abs(np.array(v_targ[1:n_bus]) - np.array(v_pred[1:n_bus]))))
        mse_bd = (sum(np.abs(np.array(t_targ[1:n_bus]) - np.array(t_pred[1:n_bus]))))
        mse_pg = sum(np.abs(np.array(pg_targ[1:n_bus] - np.array(pg_pred[1:n_bus]))))
        mse_qg = sum(np.abs(np.array(qg_targ[1:n_bus] - np.array(qg_pred[1:n_bus]))))
        plt.bar(['V [p.u.]', 'T [degree]', 'Pg [MW]', 'Qg [MVAr]'], [mse_ac, mse_bd, mse_pg, mse_qg])
        plt.xlabel('quantity')
        plt.ylabel('Mean Absolute Error (MAE)')
        plt.title(f'error total test set {powergrid}-{task}- {best_model} -{best_hp} -{best_random_seed}')
        plt.savefig(f'mse_bar_plot_{powergrid}_{task}.png')

    elif task == 'nodeopf':
        plt.figure(0)
        best_excel_file_path = models_path + powergrid + '\\' + 'summary' + powergrid + '_' + best_modelopf + '_' + task + '_'+best_hpopf+'_' + str(
        best_random_seedopf) + 's' + '.xlsx'
        workbookplot = openpyxl.load_workbook(best_excel_file_path)
        sheet = workbookplot[sheet_name_d]
        # Get data from columns A, B, C, and D
        v_targ = [cell.value for cell in sheet['A']]
        t_targ = [cell.value for cell in sheet['B']]
        pg_targ = [cell.value for cell in sheet['C']]
        qg_targ = [cell.value for cell in sheet['D']]
        v_pred = [cell.value for cell in sheet['E']]
        t_pred = [cell.value for cell in sheet['F']]
        pg_pred = [cell.value for cell in sheet['G']]
        qg_pred = [cell.value for cell in sheet['H']]

        # Calculate MSE between columns
        mse_ac = sum(np.abs(np.array(v_targ[1:n_bus] - np.array(v_pred[1:n_bus]))))
        mse_bd = sum(np.abs(np.array(t_targ[1:n_bus] - np.array(t_pred[1:n_bus]))))
        mse_pg = sum(np.abs(np.array(pg_targ[1:n_bus] - np.array(pg_pred[1:n_bus]))))
        mse_qg = sum(np.abs(np.array(qg_targ[1:n_bus] - np.array(qg_pred[1:n_bus]))))
        print(f"GRID {powergrid} v:{mse_ac},t:{mse_bd}, pg:{mse_pg},qg:{mse_qg}")

        plt.bar(['V [p.u.]', 'T [degree]', 'Pg [MW]', 'Qg [MVAr]'], [mse_ac, mse_bd, mse_pg, mse_qg])
        plt.xlabel('quantity')
        plt.ylabel('Mean Absolute Error (MAE)')
        plt.title(f'error total test set {powergrid}-{task}- {best_modelopf} -{best_hpopf} -{best_random_seedopf}')
        plt.savefig(f'mse_bar_plot_{powergrid}_{task}.png')

    if task == 'node':
        df_sheet_node = pd.DataFrame(sheet_node)
    if task == 'nodeopf':
        df_sheet_nodeopf = pd.DataFrame(sheet_nodeopf)


excel_file_path = f'processed_results_{powergrid}_modelbig.xlsx'

# Create a Pandas Excel writer using XlsxWriter as the engine
with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
    # Write each DataFrame to a different sheet
    df_sheet_node.to_excel(writer, sheet_name='node', index=False)
    df_sheet_nodeopf.to_excel(writer, sheet_name='nodeopf', index=False)
    for task in tasks:
        worksheet = writer.sheets[task]
        worksheet.insert_image('E1', f'mse_bar_plot_{powergrid}_{task}.png')

# Close the workbook when done
workbook.close()
